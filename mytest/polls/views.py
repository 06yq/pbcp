import csv
import json
import datetime
import pymongo
from io import BytesIO

from bson import ObjectId,json_util
from django.contrib.auth.hashers import make_password,check_password
from django.http import HttpResponse, JsonResponse
from django.utils.http import urlquote
from openpyxl import Workbook
from .import models
class JSONEncoder(json.JSONEncoder):
    def __init__(self, ensure_ascii=False):
        super().__init__(ensure_ascii=False)
    def default(self, o):
        if isinstance(o, ObjectId):
            return str(o)
        return json.JSONEncoder.default(self, o)
# Create your views here
#从excel中  创建表  并导入到数据库中
def creat_table(request):
    # # user = xlrd.open_workbook('C:\\Users\\Administrator\\Desktop\\user.xlsx')
    # file_path = "C:/Users/Administrator/Desktop/user.xlsx"  # 获取文件路径
    # wb = load_workbook(file_path)
    # # 写个循环遍历所有sheet
    # # sh_names = user.sheet_names
    # sh_names = wb.sheetnames
    # for sh_name in sh_names:
    #     sheet = wb[sh_name]
    #     rows = sheet.values
    #     header = next(rows) #头部
    #     for row in rows:
    #         for i,h in enumerate(header):
    #             if h == '账号':
    #                account = row[i]
    #             if h == '密码':
    #                 pwd = row[i]
    #     if sh_name == 'user':
    #         models.user.objects.create(account=account, pwd=pwd)
    #     if sh_name == 'user1':
    #         models.user1.objects.create(account=account, pwd=pwd)
    t1 = datetime.date.today()
    t2 = t1.strftime('%Y/%m/%d')
    models.gp1.objects.create(type=1,endTime=t1,startBlance=4545,startTime='2020年1月1日',qhBlance=4545,gpBlance=5656,manage=5645,bonus=12,withdrawal=225,nowMoney=545,addMoney=55,nowEquity=1.1,AddEquity=1.0,maxEquity=1,nowShow=5.2,nowBack=4,maxBack=4,gpMarket=5454,qhMarket=555545,ratio=5)
    return HttpResponse('数据添加成功')
# 注册
def register(request):
    if request.method == "POST":
        account = request.POST.get('account')
        pwd = request.POST.get('pwd')
        pwd = make_password(pwd) #密码加密
        message = ""
        type = True
        same_account = models.pbUser.objects.filter(account=account)
        if same_account:
            type = False
            message = '用户已经存在，请重新输入用户名！'
        try:
            create = models.pbUser.objects.create(account=account,pwd=pwd)
            if create:
                type = True
                message = '注册成功'
        except:
            type = False
            message = '注册失败'
        data = {'type': type, 'message': message}
        return HttpResponse(json.dumps(dict(data)))
    return HttpResponse('get请求无返回数据')
# 登陆
def login(request):
    if request.method == "POST":
        account = request.POST.get('account')
        pwd = request.POST.get('pwd')
        message=""
        type = True
        if account and pwd:  # 确保用户名和密码都不为空
            account = account.strip()
            pwd = pwd.strip()

            try:
                # list = pbUser.objects(account=account)
                user = models.pbUser.objects.get(account=account)
                print(user.account)
                #check_password 解密
                if check_password(pwd,user.pwd):
                    type = True
                    message = "登陆成功"
                    request.session['name']=user.account
                    # auth.login(request,user)
                    # return redirect('/index/',dict({'type': type, 'message': message,'account':account}))
                else:
                    type = False
                    message = "密码不正确！"
            except:
                type = False
                message = "用户不存在！"
            data = {'type': type, 'message': message}
            return HttpResponse(json.dumps(dict(data)))
    return HttpResponse('get请求无返回数据')
# 忘记密码
def forgetPwd(request):
    if request.method == "POST":
        account = request.POST.get('account', None)
        pwd = request.POST.get('pwd', None)
        message = ""
        type = True
        print(account,pwd)
        if account and pwd:  # 确保用户名和密码都不为空
            account = account.strip()
            pwd = pwd.strip()
            pwd = make_password(pwd)
            try:
                updatePwd = models.pbUser.objects.filter(account=account).first().update(pwd=pwd)
                if updatePwd:
                    message = "修改成功！"
                    type = True
                    # return redirect('/login')
            except:
                message = "用户不存在！"
                type = False
        data = {'type':type, 'message': message}
        return HttpResponse(json.dumps(dict(data)))
    return HttpResponse('get请求无返回数据')
#获取数据
def getDataList(request):
    # result = serializers.serialize("json", models.gp1.objects.all())
    name = request.GET.get('name')
    result = models.pbcpequ.objects.filter(pbcpName=name).all()
    json_list = []
    for item in result:
        json_list.append({
            "id": JSONEncoder().encode(item.id),
            "产品名": item.pbcpName,
            "日期": item.endTime,
            "初始资金": item.startBlance,
            "初始日期": item.startTime,
            "期货资产": item.qhBlance,
            "股票资产": item.gpBlance,
            "管理费": item.manage,
            "分红费": item.bonus,
            "存取金额": item.withdrawal,
            "当前权益": round(item.nowMoney, 2),
            "累计权益": round(item.addMoney, 2),
            "净值": item.nowEquity,
            "累计净值": item.AddEquity,
            "最大净值": item.maxEquity,
            "当日表现": item.nowShow,
            "回撤": item.nowBack,
            "最大回撤": item.maxBack,
            "股票占用资金": item.gpMarket,
            "期货占用资金": item.qhMarket,
            "占用比例": item.ratio,
            "tableName": name,
        })
    data = json_list
    return JsonResponse(data,charset='utf-8',safe=False)
def switchType(type):
    types = {
        '0':'股票端',
        '1':'期货端',
    }
    return types.get(type, None)
#获取净值数据
def getEquData(request):

    pbcpInfo = []
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='localhost', port=27017)
    # 2.获取数据库对象
    db = conn.shpbcp
    name = request.GET.get('name')
    my_collection = db['pbcpequ']
    count = my_collection.find({'pbcpName':name}, {'endTime': 1, '_id': 0})  # 查询某一字段的值
    time = []
    for item in count:
        time.append(item['endTime'])
    pbcpInfo.append(["time", time])
    count = my_collection.find({'pbcpName':name}, {'nowEquity': 1, '_id': 0})  # 查询某一字段的值
    equtiy = []
    for item in count:
        equtiy.append(item['nowEquity'])
        # return equtiy
    pbcpInfo.append(["equtiy", equtiy])
    count = my_collection.find({'pbcpName':name}, {'nowBack': 1, '_id': 0})  # 查询某一字段的值
    back = []
    for item in count:
        back.append(item['nowBack'])
        # return back
    count = my_collection.find({'pbcpName':name}, {'AddEquity': 1, '_id': 0})  # 查询某一字段的值
    ljEqu = []
    for item in count:
        ljEqu.append(item['AddEquity'])
    pbcpInfo.append(["back", back])
    rowsx = my_collection.find({'pbcpName':name}).limit(1)
    for row in rowsx:  # 这个循环只会执行1次
        x = row
    rows = my_collection.find({'pbcpName':name}).sort('_id', -1).limit(1)  # 倒序以后，只返回1条数据
    # infoname=['初始日期','截止日期','初始资金','净值' ,'累计净值','回撤' ,'最大回撤','当日表现','股票占用资金','期货占用资金','当前权益']
    for row in rows:  # 这个循环只会执行1次
        y = row
    pbcpInfo.append(['title',name])
    pbcpInfo.append(['startTime', x['startTime']])
    pbcpInfo.append(['startBlance', x['startBlance']])
    pbcpInfo.append(['endTime', y['endTime']])
    pbcpInfo.append(['currentEquity', (round(y['nowMoney'],2))])
    pbcpInfo.append(['nowEquity', y['nowEquity']])
    pbcpInfo.append(['AddEquity', y['AddEquity']])
    pbcpInfo.append(['maxEquity', y['maxEquity']])
    pbcpInfo.append(['nowBack', y['nowBack']])
    pbcpInfo.append(['MaxBack', y['maxBack']])
    pbcpInfo.append(['nowShow', y['nowShow']])
    pbcpInfo.append(['gpBlance', y['gpMarket']])
    pbcpInfo.append(['qhBlance', y['qhMarket']])
    pbcpInfo.append(['minEqu', min(ljEqu)])
    pbcpInfo = (dict(pbcpInfo))
    # print(pbcpInfo)
    return HttpResponse(json.dumps(pbcpInfo))
#查找最后一条数据
def getPrevData(tableName):
    import pymongo
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='localhost', port=27017)
    # 2.获取数据库对象
    db = conn.shpbcp
    name = 'pbcpequ'
    my_collection = db[name]
    rows = my_collection.find({'pbcpName':tableName}).sort('_id', -1).limit(1)  # 倒序以后，只返回1条数据
    for row in rows:  # 这个循环只会执行1次
        prevData = row
    # return HttpResponse(json_util.dumps(prevData))
    return prevData
#新增数据
def addEqu(request):
    if request.method == "POST":
        # 获取数据并计算
        tableName = request.POST.get('tableName')
        # 获取最新管理费、存取金额、分红费
        bd = []
        ma = []
        result = models.pbcpequ.objects.filter(pbcpName=tableName)
        for item in result:
            if item.bonus != 0:
                bd.append(item.bonus)
            if item.manage != 0:
                ma.append(item.manage)

        # 获取最后一条数据
        prevData = getPrevData(tableName)
        endTime = request.POST.get('endTime')
        if endTime != '':
            endTime = datetime.datetime.strptime(endTime, "%Y-%m-%d").strftime("%Y/%m/%d")
        else:
            endTime = ''
        startBlance = request.POST.get('startBlance')
        if startBlance!='':
            startBlance = float(startBlance)
        else:
            startBlance = 0
        startTime = request.POST.get('startTime')
        if startTime!='':
            startTime = datetime.datetime.strptime(startTime, "%Y-%m-%d").strftime("%Y/%m/%d")
        else:
            startTime = ''
        qhBlance = request.POST.get('qhBlance')
        if qhBlance != '':
            qhBlance = float(qhBlance)
        else:
            qhBlance = 0
        gpBlance = request.POST.get('gpBlance')
        if gpBlance != '':
            gpBlance = float(gpBlance)
        else:
            gpBlance = 0
        manage =request.POST.get('manage')
        if manage != '':
            manage = float(manage)
            manage1 = float(manage)
        elif ma:
            manage1 = ma[0]
            manage = 0
        else:
            manage1 = 0
            manage = 0
        bonus =request.POST.get('bonus')
        if bonus != '':
            bonus = float(bonus)
            bonus1 = float(bonus)
        elif bd:
            bonus1 = bd[0]
            bonus = 0
        else:
            bonus1 = 0
            bonus = 0
        withdrawal = request.POST.get('withdrawal')
        if withdrawal != '':
            withdrawal = float(withdrawal)
        else:
            withdrawal = 0
        gpMarket = request.POST.get('gpMarket')
        if gpMarket != '':
            gpMarket = float(gpMarket)
        else:
            gpMarket = 0
        qhMarket = request.POST.get('qhMarket')
        if qhMarket != '':
            qhMarket = float(qhMarket)
        else:
            qhMarket = 0
        prevData = dict(prevData)
        nowMoney = round(float(qhBlance+gpBlance),2)
        addMoney = float(nowMoney+bonus1)
        #出金入金
        nowEquity = round((((nowMoney - (withdrawal) + manage1) - (prevData['nowMoney'] + manage1)) / (prevData['nowMoney'] + manage1) + prevData['nowEquity']), 4)
        AddEquity = round((((addMoney - (withdrawal) + manage1) - (prevData['addMoney'] + manage1)) / (prevData['addMoney'] + manage1) + prevData['AddEquity']), 4)
        maxEquity = max(prevData['maxEquity'],AddEquity)
        nowShow = "%.2f%%" % (((nowEquity-prevData['nowEquity'])/prevData['nowEquity']) * 100)
        nowBack = round(((maxEquity - AddEquity)/maxEquity)*100,2)
        maxBack = max(prevData['maxBack'],nowBack)
        ratio = "%.2f%%" % (((gpMarket+qhMarket)/addMoney)*100)
        typeS = False
        message = '添加失败'
        try:
            create = models.pbcpequ.objects.create(pbcpName=tableName, endTime=endTime, startBlance=startBlance,
                                               startTime=startTime, qhBlance=qhBlance, gpBlance=gpBlance, manage=manage,
                                               bonus=bonus, withdrawal=withdrawal,nowMoney=nowMoney, addMoney=addMoney, nowEquity=nowEquity,
                                               AddEquity=AddEquity, maxEquity=maxEquity,nowShow=nowShow, nowBack=nowBack, maxBack=maxBack,
                                               gpMarket=gpMarket, qhMarket=qhMarket, ratio=ratio)
            if create:
                typeS = True
                message = '添加'+tableName+'成功'
        except:
            typeS = False
            message = '添加'+tableName+'失败'
        data = {'type': typeS, 'message': message}
        return HttpResponse(json.dumps(dict(data)))
    return HttpResponse('get请求无返回数据')
#新增产品导航
def addEquCp(request):
    # 获取数据并计算
    navCpAccount = request.POST.get('navCpAccount')
    navCpName = request.POST.get('navCpName')
    checkName = models.pbNav.objects.filter(navCpName=navCpName).all()
    if checkName:
        typeS = False
        message='产品名已存在'
    else:
        result = models.pbNav.objects.create(navCpName=navCpName,navCpAccount=navCpAccount)
        if result:
            typeS = True
            message='数据添加成功'
        else:
            typeS = False
            message = '数据添加失败'
    data = {'type':typeS,'message':message}
    return HttpResponse(json.dumps(dict(data)))
#获取产品导航
def getEquCp(request):
    result = models.pbNav.objects.all()
    json_list = []
    if result:
        for item in result:
            json_list.append({
                "id": JSONEncoder().encode(item.id),
                "产品名": item.navCpName,
            })
        data = json_list
    else:
        models.pbcpAccount.objects.none()
        json_list.append({
            "产品名": '',
        })
        data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)

#修改数据
def updateEqu(request):
    pass
#删除数据
def deleteEqu(request):
    del_id = request.POST.get('id')
    name = request.POST.get('tableName')
    print(name)
    try:
        delete = models.pbcpequ.objects.filter(id=ObjectId(del_id)).delete()
        if delete:
            typeS = True
            message = '删除成功'
    except:
        typeS = False
        message = '删除失败'
    data = {'type': typeS, 'message': message}
    return HttpResponse(json.dumps(dict(data)))

#导出所有的数据到EXCEL中(账户表)
def exPbAccountData(request):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '配邦账号表'  # 给工作表1设置标题
    row_one = ['产品名称', '证券名称', '证券账号', '证券密码','所属端', '备注']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = models.pbcpAccount.objects.all()
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj.underType=switchType(obj.underType)
        obj_info = [obj.underName, obj.bondName,obj.bondAccount,obj.bondPwd,obj.underType,obj.remarks]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    # ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '配邦账号表.xls'
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

#导出所有的数据到EXCEL中(净值表)
def exPbcpData(request):
    name = request.GET.get('name')
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '配邦净值表'  # 给工作表1设置标题
    row_one = ['日期', '初始资金', '初始日期', '期货资产','股票资产', '管理费','分红费','出金金额','当前权益','累计权益',
               '净值','累计净值','最大净值','当日表现','回撤','最大回撤','股票占用资金','期货占用资金','占用比例']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    result = models.pbcpequ.objects.filter(pbcpName=name).all()
    for obj in result:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj.nowBack = "%.2f%%" % obj.nowBack
        obj.maxBack = "%.2f%%" % obj.maxBack
        obj_info = [obj.endTime, obj.startBlance,obj.startTime,obj.qhBlance,obj.gpBlance,obj.manage,
                    obj.bonus, obj.withdrawal,obj.nowMoney,obj.addMoney,obj.nowEquity,obj.AddEquity,
                    obj.maxEquity, obj.nowShow,obj.nowBack,obj.maxBack,obj.gpMarket,obj.qhMarket,obj.ratio]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    # ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '%s净值表.xls' % name
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="配邦账号表.csv"'
    writer = csv.writer(response)
    writer.writerow(['产品名称', '证券名称', '证券账号', '证券密码', '所属端', '备注'])
    export_csvdata = models.pbcpAccount.objects.all().values_list('underId', 'bondName', 'bondAccount', 'bondPwd',
                                                                  'underType', 'remarks')
    for i in export_csvdata:
        writer.writerow(i)
    return response
#获取配邦账户
def getAccount(request):
    result = models.pbcpAccount.objects.all()
    json_list = []
    if result:
        for item in result:
            json_list.append({
                "id": JSONEncoder().encode(item.id),
                "证券名称": item.bondName,
                "证券账号": item.bondAccount,
                "证券密码": item.bondPwd,
                "所属端": item.underType,
                "所属产品": item.underId,
                "备注": item.remarks,
            })
        data = json_list
    else:
        models.pbcpAccount.objects.none()
        json_list.append({
            "证券名称":'',
            "证券账号": '',
            "证券密码": '',
            "所属端": '',
            "所属名称": '',
            "备注": '',
        })
        data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)
def getAccountId(request):
    underId = request.POST.get('underId')
    json_list = []
    try:
        result = models.pbcpAccount.objects.filter(underId=underId).all()
        if result:
            for item in result:
                json_list.append({
                    "id": JSONEncoder().encode(item.id),
                    "证券名称": item.bondName,
                    "证券账号": item.bondAccount,
                    "证券密码": item.bondPwd,
                    "所属端": item.underType,
                    "所属产品": item.underId,
                    "备注": item.remarks,
                })
            data = json_list
    except:
        json_list.append({
            "证券名称": '',
            "证券账号": '',
            "证券密码": '',
            "所属端": '',
            "所属名称": '',
            "备注": '',
        })
        data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)
def getPbcp(request):
    result = models.pbcpUnder.objects.all()
    json_list = []
    if result:
        for item in result:
            json_list.append({
                "id": JSONEncoder().encode(item.id),
                "产品ID": item.cpId,
                "产品名称": item.cpName,
                "备注": item.cpRemarks,
            })
        data = json_list
    else:
        models.pbcpAccount.objects.none()
        json_list.append({
            "产品ID":'',
            "产品名称": '',
            "备注": '',
        })
        data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)

#新增配邦账户
def addAccount(request):
    underId = request.POST.get('underId')
    underType = request.POST.get('underType')
    bondName = request.POST.get('bondName')
    bondAccount = request.POST.get('bondAccount')
    bondPwd = request.POST.get('bondPwd')
    remarks = request.POST.get('remarks')
    result = models.pbcpUnder.objects.filter(cpId=underId).all()
    if result:
        for item in result:
            print(item.cpName)
            underName = item.cpName
    else:
        underName = ''
    try:
        create = models.pbcpAccount.objects.create(underId=underId,underName=underName,underType=underType,bondName=bondName,bondAccount=bondAccount,bondPwd=bondPwd,remarks=remarks)
        if create:
            type = True
            message = '账户新增成功'
    except:
        type = False
        message = '账户新增失败'
    data = {'type':type,'message':message}
    return JsonResponse(data, charset='utf-8', safe=False)
def addPbcp(request):
    cpId = request.POST.get('cpId')
    cpName = request.POST.get('cpName')
    cpRemarks = request.POST.get('cpRemarks')
    filter = models.pbcpUnder.objects.filter(cpId=cpId)
    if filter:
        type = False
        message = '该产品Id已有，不能添加'
    else:
        try:
            create = models.pbcpUnder.objects.create(cpId=cpId,cpName=cpName,cpRemarks=cpRemarks)
            if create:
                type = True
                message = '产品新增成功'
        except:
            type = False
            message = '产品新增失败'
    data = {'type':type,'message':message}
    return JsonResponse(data, charset='utf-8', safe=False)

# 修改配邦账户
def updateAccount(request):
    id = request.POST.get('id')
    underId = request.POST.get('underId')
    underType = request.POST.get('underType')
    bondName = request.POST.get('bondName')
    bondAccount = request.POST.get('bondAccount')
    bondPwd = request.POST.get('bondPwd')
    remarks = request.POST.get('remarks')
    try:
        create = models.pbcpAccount.objects.filter(id=id).first().update(underId=underId, underType=underType, bondName=bondName,
                                                   bondAccount=bondAccount, bondPwd=bondPwd, remarks=remarks)
        if create:
            type = True
            message = '账户修改成功'
    except:
        type = False
        message = '账户修改失败'
    data = {'type': type, 'message': message}
    return JsonResponse(data, charset='utf-8', safe=False)
def updatePbcp(request):
    id = request.POST.get('id')
    cpId = request.POST.get('cpId')
    cpName = request.POST.get('cpName')
    cpRemarks = request.POST.get('cpRemarks')
    try:
        create = models.pbcpUnder.objects.filter(id=id).first().update(cpId=cpId, cpName=cpName, cpRemarks=cpRemarks)
        if create:
            type = True
            message = '产品修改成功'
    except:
        type = False
        message = '产品修改失败'
    data = {'type': type, 'message': message}
    return JsonResponse(data, charset='utf-8', safe=False)

#删除配邦账户
def deleteAccount(request):
    del_id = request.POST.get('id')
    try:
        delete = models.pbcpAccount.objects.filter(id=ObjectId(del_id)).delete()
        if delete:
            typeS = True
            message = '删除成功'
    except:
        typeS = False
        message = '删除失败'
    data = {'type': typeS, 'message': message}
    return HttpResponse(json.dumps(dict(data)))
def deletePbcp(request):
    del_id = request.POST.get('id')
    underId = request.POST.get('underId')
    filter = models.pbcpAccount.objects.filter(underId=underId)
    if filter:
        typeS = False
        message = '该产品有账号，不能删除'
    else:
        try:
            delete = models.pbcpUnder.objects.filter(id=ObjectId(del_id)).delete()
            if delete:
                typeS = True
                message = '删除成功'
        except:
            typeS = False
            message = '删除失败'
    data = {'type': typeS, 'message': message}
    return HttpResponse(json.dumps(dict(data)))

#获取资金表
def getQhfx(request):
    account = request.POST.get('account')
    startTime = request.POST.get('startTime')
    endTime = request.POST.get('endTime')
    access=0
    json_list = []
    if account != '' and startTime == '' and endTime == '':
        # account = int(account)
        accessR = models.qhfx.objects.filter(access__ne=0,timeDate__lte=endTime).all()
        result = models.qhfx.objects.filter(account=account).all()
    elif account != '' and startTime != '' or endTime != '':
        # account = int(account)
        accessR = models.qhfx.objects.filter(access__ne=0, timeDate__lte=endTime).all()
        result = models.qhfx.objects.filter(account=account, timeDate__gte=startTime, timeDate__lte=endTime).all()
    if result:
        for item in result:
            json_list.append({
                    "id": JSONEncoder().encode(item.id),
                    "账号": item.account,
                    "日期": item.timeDate,
                    "当前权益" : item.nowMoney,
                    "当日存取" : item.access,
                    "净值": item.nowEquity,
                    "最大净值": item.maxEquity,
                    "回撤": item.nowBack,
                    "最大回撤": item.maxBack,
                    "占用资金": item.market,
                    "占用比例": item.ratio,
                })
    else:
            json_list.append({
                "账号": '',
                "日期": '',
                "当前权益": '',
                "当日存取":'',
                "净值": '',
                "最大净值": '',
                "回撤": '',
                "最大回撤": '',
                "占用资金": '',
                "占用比例": '',
            })
    if accessR:
        for item in accessR:
            access+=item.access
    data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)

#获取成交记录表
def getDealRecord(request):
    account = request.POST.get('account')
    startTime = request.POST.get('startTime')
    endTime = request.POST.get('endTime')
    json_list = []
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='localhost', port=27017)
    # 2.获取数据库对象
    db = conn.shpbcp
    # name = request.GET.get('name')
    my_collection = db['dealRecord']
    if account !='' and startTime == '' and endTime =='':
        # account = int(account)
        result = my_collection.find({'account': account,'openClose':'平'})
    elif account !='' and startTime != '' or endTime !='':
        # account = int(account)
        result = my_collection.find({'account': account, "timeDate": {"$lte": endTime, "$gte": startTime},'openClose':'平'})
        # result = models.dealRecord.objects.filter(account=account,timeDate__gte=startTime,timeDate__lte=endTime).all()
    if result:
        for item in result:
            json_list.append({
                "id": JSONEncoder().encode(item['_id']),
                "账号": item['account'],
                "日期": item['timeDate'],
                "成交序号": item['dealNum'],
                "品种": item['symbol'],
                "平仓时间": item['closeTime'],
                "买/卖": item['buySell'],
                "价格": item['price'],
                "手数": item['lot'],
                "成交额": item['turnover'],
                "开/平": item['openClose'],
                "手续费": item['handling'],
                "盈亏": item['dealPL'],
                # "id": JSONEncoder().encode(item.id),
                # "账号": item.account,
                # "日期": item.timeDate,
                # "成交序号": item.dealNum,
                # "品种": item.symbol,
                # "平仓时间": item.closeTime,
                # "买/卖": item.buySell,
                # "价格": item.price,
                # "手数": item.lot,
                # "成交额": item.turnover,
                # "开/平": item.openClose,
                # "手续费": item.handling,
                # "盈亏": item.dealPL,

            })
    else:
        json_list.append({
            "账号": '',
            "日期": '',
            "成交序号": '',
            "品种": '',
            "平仓时间": '',
            "买/卖": '',
            "价格": '',
            "手数": '',
            "成交额": '',
            "开/平": '',
            "手续费": '',
            "盈亏": '',
        })
    data = json_list
    return JsonResponse(data, charset='utf-8', safe=False)
#刷新成交记录
def refershDeal(request):
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='192.168.0.160', port=27017)
    # 2.获取数据库对象
    db = conn.futures_cfmmc
    my_collection = db['TransactionDetails']

    conn1 = pymongo.MongoClient(host='localhost', port=27017)
    db1 = conn1.shpbcp
    collection = db1['dealRecord']
    collection.drop()
    result = my_collection.find({})
    if result:
        for item in result:
            collection.insert({
                "account": item['帐号'],
                "timeDate": item['实际成交日期'],
                "dealNum": item['成交序号'],
                "symbol": item['合约'],
                "closeTime": item['成交时间'],
                "buySell": item['买/卖'],
                "price": item['成交价'],
                "lot": item['手数'],
                "turnover": item['成交额'],
                "openClose": item['开/平'],
                "handling": item['手续费'],
                "dealPL": item['平仓盈亏'],
            })
        Type=True
        message='数据刷新成功'
    else:
        Type = False
        message = '数据刷新失败'
    data = {'type': Type, 'message': message}
    return HttpResponse(json.dumps(dict(data)))

#获取期货账号
def getQhAccount(request):
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='192.168.0.160', port=27017)
    # 2.获取数据库对象
    db = conn.futures_cfmmc
    my_collection = db['account']
    result = my_collection.find({})
    json_data = []
    if result:
        for item in result:
            json_data.append({
                '证券账号':item['user_id'],
                '证券名称':item['notes']
            })
    return JsonResponse(json_data, charset='utf-8', safe=False)
#查找最后一条数据(ZJ)
def getPrevDataZJ(account):
    import pymongo
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='localhost', port=27017)
    # 2.获取数据库对象
    db = conn.shpbcp
    my_collection = db['qhfx']
    rows = my_collection.find({'account':account}).sort('_id', -1).limit(1)  # 倒序以后，只返回1条数据
    rows1 = my_collection.find({'account':account}).limit(1)  #只返回1条数据
    data=[]
    for row in rows:  # 这个循环只会执行1次
        prevData = row
    for row in rows1:  # 这个循环只会执行1次
        prevData1 = row
    data.append({'lastData':prevData,'firstData':prevData1})
    # return HttpResponse(json_util.dumps(prevData))
    return data
#刷新资金记录
def refershMoney(request):
    account = request.POST.get('account')
    nowDate = request.POST.get('nowDate')
    endTime = datetime.datetime.strptime(nowDate, "%Y-%m-%d").strftime("%Y/%m/%d")
    preData = getPrevDataZJ(account)
    result=[]
    for item in preData:
        firstMoney = item['firstData']['nowMoney']
        lastMaxEquity = item['lastData']['maxEquity']
        lastMaxBack = item['lastData']['maxBack']
    resultZJ = models.qhfx.objects.filter(account=account, timeDate=nowDate).all()
    navCpAccount = models.pbNav.objects.filter(navCpAccount=account).all()
    if navCpAccount:
        for item in navCpAccount:
            pbcpName = item.navCpName
    else:
        pbcpName=''
    if resultZJ:
        message = '已是最新数据'
    else:
        result = models.pbcpequ.objects.filter(pbcpName=pbcpName, endTime=endTime).all()
    if result:
        for item in result:
            nowMoney = round((item.nowMoney), 2)
            market = item.qhMarket
            # firstMoney = preData['firstData']['nowMoney']
            nowEquity = round((nowMoney / firstMoney), 4)
            maxEquity = max(nowEquity, lastMaxEquity)
            nowBack = round(((maxEquity - nowEquity) / maxEquity) * 100, 2)
            maxBack = max(nowBack, lastMaxBack)
            ratio = "%.2f%%" % ((market / nowMoney) * 100)
        models.qhfx.objects.create(account=account, timeDate=nowDate, access=0, nowMoney=nowMoney,
                                   nowEquity=nowEquity, maxEquity=maxEquity, nowBack=nowBack, maxBack=maxBack,
                                   market=market, ratio=ratio)
        message = '数据更新'
    else:
        message = '今天最新数据没有更新,请稍后再试'
    data = {'message':message}
    return JsonResponse(data, charset='utf-8', safe=False)

#导出数据(ZJ)
def exportZJData(request):
    account = request.GET.get('account')
    startTime = request.GET.get('startTime')
    endTime = request.GET.get('endTime')
    print(startTime, endTime)
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '配邦净值表'  # 给工作表1设置标题
    row_one = ['账号','日期','当前权益','净值','最大净值', '回撤', '最大回撤', '占用资金', '占用比例']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    result = models.qhfx.objects.filter(account=account,timeDate__gte=startTime, timeDate__lte=endTime).all()
    for obj in result:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj.nowBack = "%.2f%%" % obj.nowBack
        obj.maxBack = "%.2f%%" % obj.maxBack
        obj_info = [obj.account,obj.timeDate, obj.nowMoney,obj.nowEquity,obj.maxEquity, obj.nowBack, obj.maxBack, obj.market, obj.ratio]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    # ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '%s净值表.xls' % account
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

def exportDealData(request):
    account = request.GET.get('account')
    startTime = request.GET.get('startTime')
    endTime = request.GET.get('endTime')
    # 1.连接数据库服务器,获取客户端对象
    conn = pymongo.MongoClient(host='localhost', port=27017)
    # 2.获取数据库对象
    db = conn.shpbcp
    my_collection = db['dealRecord']
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '成交记录表'  # 给工作表1设置标题
    row_one = ['账号', '日期', '成交序号', '品种', '平仓时间', '买/卖', '价格', '手数', '成交额','开/平','手续费','盈亏']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    result = my_collection.find({'account': account, "timeDate": {"$lte": endTime, "$gte": startTime}, 'openClose': '平'})
    for obj in result:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj['account'], obj['timeDate'], obj['dealNum'], obj['symbol'], obj['closeTime'], obj['buySell'], obj['price'],
                    obj['lot'], obj['turnover'],obj['openClose'],obj['handling'],obj['dealPL']]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    # ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = '%s成交记录表.xls' % account
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response